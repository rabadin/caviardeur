from pathlib import Path

import openpyxl

from ..readers.base import DocumentContent


def write_xlsx(content: DocumentContent, output_path: Path, source_path: Path | None = None) -> None:
    """Write pseudonymized content to an .xlsx file.

    If source_path is provided, opens the original workbook to preserve formatting.
    Otherwise, creates a new workbook (used for .xls → .xlsx conversion).
    """
    original_format = content.metadata.get("format", "xlsx")

    if source_path and original_format == "xlsx":
        wb = openpyxl.load_workbook(str(source_path))
    else:
        wb = openpyxl.Workbook()
        # Remove the default sheet — we'll create sheets as needed
        if wb.worksheets:
            wb.remove(wb.active)

    # Build lookup: (sheet, row, col) -> new text
    cell_map: dict[tuple[str, int, int], str] = {}
    sheets_seen: set[str] = set()
    for chunk in content.chunks:
        loc = chunk.location
        loc_type = loc.get("type", "")
        if loc_type in ("xlsx_cell", "xls_cell"):
            sheet_name = loc["sheet"]
            row = loc["row"]
            col = loc["col"]
            cell_map[(sheet_name, row, col)] = chunk.text
            sheets_seen.add(sheet_name)

    if original_format == "xls":
        # For xls conversion, create sheets and populate
        for sheet_name in sorted(sheets_seen):
            ws = wb.create_sheet(title=sheet_name)
            for (s, r, c), text in cell_map.items():
                if s == sheet_name:
                    ws.cell(row=r, column=c, value=text)
    else:
        # For xlsx, update cells in place
        for (sheet_name, row, col), text in cell_map.items():
            if sheet_name in wb.sheetnames:
                wb[sheet_name].cell(row=row, column=col, value=text)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
