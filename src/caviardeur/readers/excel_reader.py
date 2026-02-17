from pathlib import Path

import openpyxl
import xlrd

from .base import DocumentContent, TextChunk


def read_xlsx(path: Path) -> DocumentContent:
    """Read an .xlsx file, extracting text at the cell level."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    chunks: list[TextChunk] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value is not None:
                    text = str(cell.value)
                    if text.strip():
                        chunks.append(
                            TextChunk(
                                text=text,
                                location={
                                    "type": "xlsx_cell",
                                    "sheet": sheet_name,
                                    "row": row_idx,
                                    "col": col_idx,
                                    "coordinate": cell.coordinate,
                                },
                            )
                        )
                        chunks.append(
                            TextChunk(
                                text=" ",
                                location={"type": "xlsx_cell_separator"},
                            )
                        )

    content = DocumentContent(chunks=chunks, metadata={"source_path": str(path), "format": "xlsx"})
    content.assign_offsets()
    return content


def read_xls(path: Path) -> DocumentContent:
    """Read an .xls file using xlrd. Output will be written as .xlsx."""
    wb = xlrd.open_workbook(str(path))
    chunks: list[TextChunk] = []

    for sheet_idx in range(wb.nsheets):
        sheet = wb.sheet_by_index(sheet_idx)
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                if cell.value and str(cell.value).strip():
                    text = str(cell.value)
                    chunks.append(
                        TextChunk(
                            text=text,
                            location={
                                "type": "xls_cell",
                                "sheet": sheet.name,
                                "row": row_idx + 1,
                                "col": col_idx + 1,
                            },
                        )
                    )
                    chunks.append(
                        TextChunk(
                            text=" ",
                            location={"type": "xls_cell_separator"},
                        )
                    )

    content = DocumentContent(
        chunks=chunks,
        metadata={"source_path": str(path), "format": "xls"},
    )
    content.assign_offsets()
    return content
